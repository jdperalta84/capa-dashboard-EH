"""
CAPA KPI Metrics — Streamlit Dashboard
=======================================
Run with:  streamlit run streamlit_app.py
"""

import streamlit as st
import pandas as pd
import io
import warnings
import xlrd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import altair as alt

warnings.filterwarnings("ignore", category=UserWarning)

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="CAPA KPI Dashboard", layout="wide")

TODAY = date.today()
OPEN_THRESHOLD_DAYS = 90

# ── Excel styling palette (shared by both report builders) ────────────────────
NAVY = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "DCE6F1"
ALT_ROW = "EEF3FA"
WHITE = "FFFFFF"
DARK_TEXT = "1F3864"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
RED_BG = "FCE4D6"


def navy_fill():  return PatternFill("solid", fgColor=NAVY)
def blue_fill():  return PatternFill("solid", fgColor=MED_BLUE)
def alt_fill():   return PatternFill("solid", fgColor=ALT_ROW)
def white_fill(): return PatternFill("solid", fgColor=WHITE)
def gray_fill():  return PatternFill("solid", fgColor=LIGHT_GRAY)


def hdr_font(sz=11):  return Font(name="Arial", bold=True, color=WHITE, size=sz)
def body_font(sz=11): return Font(name="Arial", size=sz, color=DARK_TEXT)
def bold_font(sz=11): return Font(name="Arial", bold=True, size=sz, color=DARK_TEXT)


def thin_border():
    s = Side(style="thin", color=MED_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)


def xcenter(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def xleft(wrap=False):   return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def autofit(ws, max_w=40, min_w=10):
    for col_cells in ws.columns:
        length = min_w
        for cell in col_cells:
            if cell.value:
                length = max(length, min(max_w, len(str(cell.value)) + 3))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length


def _xlrd_sheet_to_df(wb, sheet_name):
    """Read a sheet from an xlrd.Book directly into a DataFrame.

    Pandas 2.3+ no longer accepts an xlrd.Book object as the io argument to
    pd.read_excel, so we materialize the sheet ourselves. Excel date cells
    are converted to pandas datetimes; empty cells become None.
    """
    sheet = wb.sheet_by_name(sheet_name)
    if sheet.nrows == 0:
        return pd.DataFrame()
    headers = [str(v).strip() if v is not None else "" for v in sheet.row_values(0)]
    rows = []
    for r in range(1, sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            val = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    val = xlrd.xldate.xldate_as_datetime(val, wb.datemode)
                except Exception:
                    pass
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                val = bool(val)
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                val = None
            row.append(val)
        rows.append(row)
    return pd.DataFrame(rows, columns=headers)


# ── Data loading ──────────────────────────────────────────────────────────────
def load_data(uploaded_files):
    """Parse uploaded export_CAPA *.xls / *.xlsx files and return a merged DataFrame.
    Shows a progress bar while processing multiple files.
    Handles both legacy format (Capas/Taken sheets) and 2026+ format (CAR/PAR/CAF/PTO sheets).
    """
    # Progress bar (Streamlit UI)
    progress = st.progress(0)
    total = len(uploaded_files)

    capas_frames = []
    for i, uploaded in enumerate(uploaded_files, 1):
        # Derive location name from filename
        location = (
            uploaded.name
            .replace("export_CAPA ", "")
            .replace(".xlsx", "")
            .replace(".xls", "")
            .strip()
        )

        raw_bytes = uploaded.read()
        uploaded.seek(0)  # reset for any later use

        # Try to detect file format and open with appropriate engine
        wb = None
        is_legacy_format = False
        try:
            # First try xlrd for old .xls (BIFF) format
            wb = xlrd.open_workbook(
                file_contents=raw_bytes,
                ignore_workbook_corruption=True,
            )
            is_legacy_format = True
        except Exception:
            pass

        if wb is None:
            # Try openpyxl for .xlsx format (files with .xls extension but xlsx content)
            try:
                import openpyxl
                # Write bytes to temp file for openpyxl
                import tempfile
                import os
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(raw_bytes)
                    tmp_path = tmp.name
                try:
                    wb = openpyxl.load_workbook(tmp_path, data_only=True)
                finally:
                    os.unlink(tmp_path)
            except Exception as e:
                st.error(f"Failed to open workbook for {uploaded.name}: {e}")
                continue

        try:
            if is_legacy_format:
                # Legacy format: Capas and Taken sheets
                capas = _xlrd_sheet_to_df(wb, "Capas")
                # Keep only selected CAPA types
                _include = [
                    "Client Complaint", "Client complaint",
                    "Customer Complaint", "Customer complaint",
                    "Site complaint",
                    "PT Outlier",
                    "Proficiency Testing Outlier",
                    "Proficiency test and Round Robins",
                    "Internal Audit",
                ]
                if "Type" in capas.columns:
                    mask = capas["Type"].astype(str).str.strip().str.lower().isin([e.lower() for e in _include])
                    capas = capas[mask]

                capas["Location"] = location
                capas["Date of notification"] = pd.to_datetime(
                    capas["Date of notification"], dayfirst=True, errors="coerce"
                )
                capas["Date closed"] = pd.to_datetime(
                    capas["Date closed"], dayfirst=True, errors="coerce"
                )

                try:
                    taken = _xlrd_sheet_to_df(wb, "Taken")
                except Exception as e:
                    st.error(f"Failed to read 'Taken' sheet: {e}")
                    continue
                taken["Date of completion"] = pd.to_datetime(
                    taken["Date of completion"], dayfirst=True, errors="coerce"
                )

                task_groups = taken.groupby("Number")

                def resolve_closed_date(row, _tg=task_groups):
                    num = row["Number"]
                    capas_date = row["Date closed"]
                    if num in _tg.groups:
                        group = _tg.get_group(num)
                        completed = group[group["Completed"].str.strip().str.lower() == "yes"]
                        max_date = completed["Date of completion"].dropna().max()
                        if pd.notna(max_date):
                            return max_date
                    return capas_date

                capas["Effective closed date"] = capas.apply(resolve_closed_date, axis=1)
                capas["Effective closed date"] = pd.to_datetime(capas["Effective closed date"], errors="coerce")
                capas_frames.append(capas)
            else:
                # 2026+ format: CAR, PAR, CAF, PTO sheets with location in first column
                # Load CAR sheet (Corrective Action Reports - main CAPA source)
                car_sheet_name = None
                for name in wb.sheetnames:
                    if "CAR" in name.upper():
                        car_sheet_name = name
                        break

                if car_sheet_name is None:
                    st.error(f"No CAR sheet found in {uploaded.name}")
                    continue

                # Read CAR sheet with openpyxl
                # Note: Row 1 = English headers, Row 2 = Spanish translations, Row 3+ = data
                car_data = []
                ws = wb[car_sheet_name]
                headers = None
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if row_idx == 1:
                        raw = [str(h).strip() if h is not None else "" for h in row]
                        # Make headers unique and non-empty so DataFrame construction
                        # and later column lookups don't collide or fail.
                        seen = {}
                        headers = []
                        for h in raw:
                            base = h or "_unnamed"
                            if base in seen:
                                seen[base] += 1
                                headers.append(f"{base}__{seen[base]}")
                            else:
                                seen[base] = 0
                                headers.append(base)
                        continue
                    if row_idx == 2:
                        # Skip Spanish translation row
                        continue

                    # Extract location from first column (it's dynamic based on row)
                    loc_value = row[0] if len(row) > 0 else None
                    # Skip rows where first column looks like a header translation
                    if loc_value and isinstance(loc_value, str) and loc_value.strip():
                        if "locación" in loc_value.lower() or "id de la locación" in loc_value.lower():
                            continue  # Skip translation row
                        car_data.append(row)

                if not car_data:
                    st.error(f"No CAR data found in {uploaded.name}")
                    continue

                # Build DataFrame with normalized column names
                capas = pd.DataFrame(car_data, columns=headers)

                # Normalize column names - find the key columns by pattern matching.
                # Patterns are intentionally permissive so common header variants
                # ("Date Closed", "Closed Date", "Notification Date", etc.) all map.
                col_map = {}
                for col in capas.columns:
                    col_lower = col.lower().strip()
                    if not col_lower or col_lower.startswith("_unnamed"):
                        continue
                    if "location" in col_lower and "id" not in col_lower:
                        col_map[col] = "Location"
                    elif (
                        "initialized" in col_lower
                        or "notification" in col_lower
                        or "open date" in col_lower
                        or "date opened" in col_lower
                    ):
                        col_map[col] = "Date of notification"
                    elif "closed" in col_lower and ("date" in col_lower or "effective" in col_lower):
                        col_map[col] = "Date closed"
                    elif "complete corrective" in col_lower or "approved date" in col_lower:
                        col_map[col] = "Complete corrective actions (Approved Date)"
                    elif "car #" in col_lower or "car number" in col_lower or col_lower in ("number", "id", "car id"):
                        col_map[col] = "Number"
                    elif "car type" in col_lower or col_lower == "type":
                        col_map[col] = "Type"
                    elif "brief description" in col_lower or ("nc" in col_lower and "description" in col_lower):
                        col_map[col] = "Description"
                    elif "status" in col_lower:
                        col_map[col] = "Status"

                capas = capas.rename(columns=col_map)

                # Ensure every column we reference downstream actually exists,
                # so a missing/oddly-named source column never raises KeyError.
                for required in ("Location", "Number", "Type",
                                 "Date of notification", "Date closed"):
                    if required not in capas.columns:
                        capas[required] = pd.NA

                # Extract location from the first data column if not already mapped
                if capas["Location"].isna().all() and capas.shape[1] > 0:
                    capas["Location"] = capas.iloc[:, 0]

                # Filter to relevant CAPA types
                _include = [
                    "Client Complaint", "Client complaint",
                    "Customer Complaint", "Customer complaint",
                    "Site complaint",
                    "PT Outlier",
                    "Proficiency Testing Outlier",
                    "Proficiency test and Round Robins",
                    "Internal Audit",
                    "Complaint",
                    "Other (QC, customer alert, improper procedure)",
                    "Customer audit",
                    "3rd party audit",
                    "Internal assessment",
                ]
                mask = capas["Type"].astype(str).str.strip().str.lower().isin(
                    [e.lower() for e in _include]
                )
                if mask.any():
                    capas = capas[mask]
                # If no Type matches at all, keep all rows so the user still
                # sees their data instead of a silently empty dashboard.

                # Parse dates
                capas["Date of notification"] = pd.to_datetime(
                    capas["Date of notification"], errors="coerce"
                )
                capas["Date closed"] = pd.to_datetime(
                    capas["Date closed"], errors="coerce"
                )

                # Use Complete corrective actions date as fallback for effective closed date
                if "Complete corrective actions (Approved Date)" in capas.columns:
                    capas["Effective closed date"] = capas["Date closed"].fillna(
                        pd.to_datetime(capas["Complete corrective actions (Approved Date)"], errors="coerce")
                    )
                else:
                    capas["Effective closed date"] = capas["Date closed"]

                # Add Status column if not present
                if "Status" not in capas.columns:
                    capas["Status"] = capas["Date closed"].apply(
                        lambda x: "Closed" if pd.notna(x) else "Open"
                    )

                # Clean up Location: coerce to str, blank/NaN -> "Unknown" so
                # sorting and multiselect can't trip on mixed types.
                capas["Location"] = (
                    capas["Location"]
                    .astype("object")
                    .where(capas["Location"].notna(), "Unknown")
                    .astype(str)
                    .str.strip()
                    .replace("", "Unknown")
                )

                capas_frames.append(capas)

        except Exception as e:
            st.error(f"Failed to process {uploaded.name}: {e}")
            continue

        # Update progress bar
        progress.progress(i / total)

    progress.empty()
    if not capas_frames:
        st.error("No CAPA data could be loaded. Please check your files.")
        return pd.DataFrame(), []
    all_capas = pd.concat(capas_frames, ignore_index=True)
    locations = sorted(
        str(loc) for loc in all_capas["Location"].dropna().unique()
        if str(loc).strip()
    )
    return all_capas, locations


# ── Metric calculation ────────────────────────────────────────────────────────
def compute_metrics(all_capas, method):
    """
    method: "official" uses Date closed only
            "taskdates" uses Effective closed date
    Returns dict with all KPIs, location DataFrame, and detail DataFrames.
    """
    if method == "official":
        closed_col = "Date closed"
    else:
        closed_col = "Effective closed date"

    is_closed = all_capas[closed_col].notna()
    is_open = ~is_closed

    closed_2025 = is_closed & (all_capas[closed_col].dt.year == 2025)
    closed_2026 = is_closed & (all_capas[closed_col].dt.year == 2026)

    days_open = (pd.Timestamp(TODAY) - all_capas["Date of notification"]).dt.days
    open_gt90 = is_open & (days_open > OPEN_THRESHOLD_DAYS)

    def avg_days(mask):
        df = all_capas[mask].copy()
        df["days_to_close"] = (df[closed_col] - df["Date of notification"]).dt.days
        vals = df["days_to_close"].dropna()
        return vals.mean() if not vals.empty else float("nan")

    avg_2025 = avg_days(closed_2025)
    avg_2026 = avg_days(closed_2026)

    # Per-location breakdown
    location_rows = []
    for loc in sorted(all_capas["Location"].unique()):
        mask = all_capas["Location"] == loc
        loc_avg_2025 = avg_days(mask & closed_2025)
        loc_avg_2026 = avg_days(mask & closed_2026)
        location_rows.append({
            "Location": loc,
            "Avg Days to Close (2025)": round(loc_avg_2025, 1) if pd.notna(loc_avg_2025) else "N/A",
            "Closed 2025": int((mask & closed_2025).sum()),
            "Avg Days to Close (2026)": round(loc_avg_2026, 1) if pd.notna(loc_avg_2026) else "N/A",
            "Closed 2026": int((mask & closed_2026).sum()),
            "Open": int((mask & is_open).sum()),
            f"Open > {OPEN_THRESHOLD_DAYS} days": int((mask & open_gt90).sum()),
        })

    df_location = pd.DataFrame(location_rows)

    # Detail DataFrames
    cols_detail = ["Location", "Number", "Date of notification",
                   "Date closed", "Effective closed date", "Status"]
    details = {
        "Closed 2025": all_capas[closed_2025][cols_detail].copy(),
        "Closed 2026": all_capas[closed_2026][cols_detail].copy(),
        "Open": all_capas[is_open][cols_detail].copy(),
        f"Open > {OPEN_THRESHOLD_DAYS} days": all_capas[open_gt90][cols_detail].copy(),
    }

    return {
        "avg_2025": avg_2025,
        "avg_2026": avg_2026,
        "closed_2025": int(closed_2025.sum()),
        "closed_2026": int(closed_2026.sum()),
        "open": int(is_open.sum()),
        "open_gt90": int(open_gt90.sum()),
        "df_location": df_location,
        "details": details,
    }


# ── Excel report builder ─────────────────────────────────────────────────────
def build_excel_report(metrics, method):
    """Build the styled Excel workbook in memory and return bytes."""
    m = metrics
    df_location = m["df_location"]
    avg_days_to_close_2025 = m["avg_2025"]
    avg_days_to_close_2026 = m["avg_2026"]
    total_closed_2025 = m["closed_2025"]
    total_closed_2026 = m["closed_2026"]
    total_open = m["open"]
    total_open_gt90 = m["open_gt90"]

    is_taskdates = method == "taskdates"
    title_suffix = " — Task Date Priority" if is_taskdates else ""

    wb = Workbook()
    wb.remove(wb.active)

    # ── Dashboard sheet ───────────────────────────────────────────────────────
    dash = wb.create_sheet("Dashboard")
    dash.sheet_view.showGridLines = False

    dash.merge_cells("B2:I2")
    dash["B2"].value = f"CAPA KPI Dashboard{title_suffix}"
    dash["B2"].font = Font(name="Arial", bold=True, size=22, color=NAVY)
    dash["B2"].alignment = xleft()

    dash.merge_cells("B3:I3")
    dash["B3"].value = f"Report date: {TODAY.strftime('%B %d, %Y')}"
    dash["B3"].font = Font(name="Arial", size=11, color="595959")
    dash["B3"].alignment = xleft()

    dash.row_dimensions[2].height = 36
    dash.row_dimensions[3].height = 18

    def kpi_card(ws, label_row, val_row, col_start, col_end, label, value, bg_val=LIGHT_BLUE):
        col_s = get_column_letter(col_start)
        col_e = get_column_letter(col_end)
        ws.merge_cells(f"{col_s}{label_row}:{col_e}{label_row}")
        lc = ws[f"{col_s}{label_row}"]
        lc.value = label
        lc.fill = navy_fill()
        lc.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        lc.alignment = xcenter()
        ws.row_dimensions[label_row].height = 22
        ws.merge_cells(f"{col_s}{val_row}:{col_e}{val_row}")
        vc = ws[f"{col_s}{val_row}"]
        vc.value = value
        vc.fill = PatternFill("solid", fgColor=bg_val)
        vc.font = Font(name="Arial", bold=True, size=26, color=NAVY)
        vc.alignment = xcenter()
        ws.row_dimensions[val_row].height = 44
        for r in [label_row, val_row]:
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).border = thin_border()

    cards = [
        ("Avg Days to Close (2025)", f"{avg_days_to_close_2025:.1f} days"),
        ("Total Closed in 2025", f"{total_closed_2025:,}"),
        ("Avg Days to Close (2026)", f"{avg_days_to_close_2026:.1f} days"),
        ("Total Closed in 2026", f"{total_closed_2026:,}"),
        ("Currently Open", f"{total_open:,}"),
        (f"Open > {OPEN_THRESHOLD_DAYS} Days", f"{total_open_gt90:,}"),
    ]
    card_positions = [
        (5, 6, 2, 5), (5, 6, 6, 9),
        (8, 9, 2, 5), (8, 9, 6, 9),
        (11, 12, 2, 5), (11, 12, 6, 9),
    ]
    card_colors = [LIGHT_BLUE, LIGHT_BLUE, LIGHT_BLUE, LIGHT_BLUE, RED_BG, RED_BG]

    for i, (label, value) in enumerate(cards):
        lr, vr, cs, ce = card_positions[i]
        kpi_card(dash, lr, vr, cs, ce, label, value, bg_val=card_colors[i])

    # Location table on dashboard
    tbl_start = 15
    dash.merge_cells(f"B{tbl_start}:H{tbl_start}")
    dash[f"B{tbl_start}"].value = "Performance by Location"
    dash[f"B{tbl_start}"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    dash[f"B{tbl_start}"].fill = navy_fill()
    dash[f"B{tbl_start}"].alignment = xleft()
    dash.row_dimensions[tbl_start].height = 22

    loc_cols = ["Location", "Avg Days to Close (2025)", "Closed 2025",
                "Avg Days to Close (2026)", "Closed 2026", "Open", f"Open > {OPEN_THRESHOLD_DAYS} days"]

    hdr_row = tbl_start + 1
    for ci, col in enumerate(loc_cols, start=2):
        cell = dash.cell(row=hdr_row, column=ci)
        cell.value = col
        cell.fill = blue_fill()
        cell.font = hdr_font(10)
        cell.border = thin_border()
        cell.alignment = xcenter(wrap=True)
    dash.row_dimensions[hdr_row].height = 30

    for ri, (_, row_data) in enumerate(df_location.iterrows(), start=hdr_row + 1):
        fill = white_fill() if ri % 2 == 0 else alt_fill()
        values = [row_data[c] for c in loc_cols]
        for ci, val in enumerate(values, start=2):
            cell = dash.cell(row=ri, column=ci)
            cell.value = val
            cell.fill = fill
            cell.font = body_font(10)
            cell.border = thin_border()
            cell.alignment = xleft() if ci == 2 else xcenter()
        dash.row_dimensions[ri].height = 18

    col_widths = {2: 16, 3: 22, 4: 14, 5: 22, 6: 14, 7: 10, 8: 14, 9: 3}
    for col, w in col_widths.items():
        set_col_width(dash, col, w)
    set_col_width(dash, 1, 3)
    dash.freeze_panes = "B5"

    # ── By Location sheet ─────────────────────────────────────────────────────
    loc_ws = wb.create_sheet("By Location")
    loc_ws.sheet_view.showGridLines = False

    loc_headers = ["Location", "Avg Days to Close (2025)", "Closed 2025",
                   "Avg Days to Close (2026)", "Closed 2026", "Open", f"Open > {OPEN_THRESHOLD_DAYS} Days"]
    for ci, h in enumerate(loc_headers, start=1):
        cell = loc_ws.cell(row=1, column=ci)
        cell.value = h
        cell.fill = navy_fill()
        cell.font = hdr_font()
        cell.border = thin_border()
        cell.alignment = xcenter(wrap=True)
    loc_ws.row_dimensions[1].height = 30

    for ri, (_, row_data) in enumerate(df_location.iterrows(), start=2):
        fill = white_fill() if ri % 2 == 0 else alt_fill()
        values = [row_data[c] for c in loc_cols]
        for ci, val in enumerate(values, start=1):
            cell = loc_ws.cell(row=ri, column=ci)
            cell.value = val
            cell.fill = fill
            cell.font = body_font()
            cell.border = thin_border()
            cell.alignment = xleft() if ci == 1 else xcenter()
        loc_ws.row_dimensions[ri].height = 18
    autofit(loc_ws, max_w=28)
    loc_ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    sum_ws = wb.create_sheet("Summary")
    sum_ws.sheet_view.showGridLines = False

    for ci, h in enumerate(["Metric", "Value"], start=1):
        cell = sum_ws.cell(row=1, column=ci)
        cell.value = h
        cell.fill = navy_fill()
        cell.font = hdr_font()
        cell.border = thin_border()
        cell.alignment = xcenter()
    sum_ws.row_dimensions[1].height = 24

    summary_rows = [
        ("Avg Days to Close (2025)", f"{avg_days_to_close_2025:.1f}"),
        ("Total Closed in 2025", total_closed_2025),
        ("Avg Days to Close (2026)", f"{avg_days_to_close_2026:.1f}"),
        ("Total Closed in 2026", total_closed_2026),
        ("Currently Open", total_open),
        (f"Open > {OPEN_THRESHOLD_DAYS} Days", total_open_gt90),
    ]
    for ri, (metric, value) in enumerate(summary_rows, start=2):
        fill = white_fill() if ri % 2 == 0 else alt_fill()
        for ci, val in enumerate([metric, value], start=1):
            cell = sum_ws.cell(row=ri, column=ci)
            cell.value = val
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = xleft() if ci == 1 else xcenter()
            cell.font = body_font() if ci == 1 else bold_font()
        sum_ws.row_dimensions[ri].height = 20
    sum_ws.column_dimensions["A"].width = 36
    sum_ws.column_dimensions["B"].width = 16
    sum_ws.freeze_panes = "A2"

    # ── Logic Notes sheet ─────────────────────────────────────────────────────
    log_ws = wb.create_sheet("Logic Notes")
    log_ws.sheet_view.showGridLines = False

    log_ws.merge_cells("A1:B1")
    log_ws["A1"].value = "CAPA Metrics — Logic & Methodology"
    log_ws["A1"].font = Font(name="Arial", bold=True, size=16, color=WHITE)
    log_ws["A1"].fill = navy_fill()
    log_ws["A1"].alignment = xcenter()
    log_ws.row_dimensions[1].height = 32

    if is_taskdates:
        logic_content = [
            ("ALTERNATIVE METHOD: TASK DATE PRIORITY", "", True),
            ("How this differs from the official method",
             "The official report uses only the 'Date closed' field on the main Capas tab. "
             "This version prioritises task completion dates from the Taken sheet, falling back "
             "to the Capas date if no completed tasks exist.", False),
            ("Why task dates may be considered",
             "Task completion dates reflect when the actual corrective work was finished.", False),
            ("Known trade-off: shorter averages",
             "Using task dates produces lower avg days figures.", False),
            ("Known trade-off: more CAPAs counted as closed",
             "CAPAs with completed tasks but no Date closed on the Capas tab are counted as closed here.", False),
            ("CAPA type filtering",
             "Only CAPAs whose 'Type' matches the selected categories are retained.", False),
        ]
    else:
        logic_content = [
            ("WHY WE USE DATE CLOSED ON THE MAIN TAB — NOT TASK DATES", "", True),
            ("The 'Date closed' on the Capas tab is the official system sign-off",
             "Someone deliberately marked the CAPA closed on that date.", False),
            ("Task completion dates are inputs, not closure",
             "A CAPA may have many tasks. The last task finishing does not mean the CAPA is "
             "formally reviewed and closed.", False),
            ("Consistency and auditability",
             "The Capas tab Date closed is a single, unambiguous field.", False),
            ("CAPA type filtering",
                 "Only CAPAs whose 'Type' matches the selected categories are retained.", False),
        ]

    logic_content += [
        ("HOW EACH METRIC IS CALCULATED", "", True),
        ("Avg days to close (2025)",
         "All CAPAs with a closed date in 2025. Formula: closed date - Date of notification, averaged.", False),
        ("Avg days to close (2026)", "Same logic for 2026.", False),
        ("Currently open", "CAPAs with no closed date.", False),
        ("Open > 90 days",
         f"Open CAPAs where today - Date of notification > {OPEN_THRESHOLD_DAYS} days.", False),
        ("DATA SOURCES", "", True),
        ("Input files", "All files matching 'export_CAPA *.xls' or 'export_CAPA *.xlsx' in the project folder.", False),
        ("Sheets used", "'Capas' sheet + 'Taken' sheet.", False),
    ]

    row = 2
    for rule, detail, is_section in logic_content:
        if is_section:
            log_ws.merge_cells(f"A{row}:B{row}")
            cell = log_ws[f"A{row}"]
            cell.value = rule
            cell.fill = blue_fill()
            cell.font = Font(name="Arial", bold=True, size=11, color=WHITE)
            cell.alignment = xleft()
            cell.border = thin_border()
            log_ws.row_dimensions[row].height = 22
        else:
            fill = gray_fill() if row % 2 == 0 else white_fill()
            for ci, val in enumerate([rule, detail], start=1):
                cell = log_ws.cell(row=row, column=ci)
                cell.value = val
                cell.fill = fill
                cell.font = bold_font(10) if ci == 1 else body_font(10)
                cell.alignment = xleft(wrap=True)
                cell.border = thin_border()
            log_ws.row_dimensions[row].height = 42
        row += 1
    log_ws.column_dimensions["A"].width = 45
    log_ws.column_dimensions["B"].width = 80

    # ── Detail sheets ─────────────────────────────────────────────────────────
    cols_to_show = ["Location", "Number", "Date of notification",
                    "Date closed", "Effective closed date", "Status"]
    date_cols_set = {"Date of notification", "Date closed", "Effective closed date"}
    DATE_FMT = "YYYY-MM-DD"

    for sheet_name, df_detail in m["details"].items():
        ws = wb.create_sheet(f"{sheet_name} Detail")
        ws.sheet_view.showGridLines = False

        for ci, col in enumerate(cols_to_show, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.value = col
            cell.fill = navy_fill()
            cell.font = hdr_font()
            cell.border = thin_border()
            cell.alignment = xcenter(wrap=True)
        ws.row_dimensions[1].height = 24

        df_out = df_detail[cols_to_show].copy()
        for ri, (_, row_vals) in enumerate(df_out.iterrows(), start=2):
            fill = white_fill() if ri % 2 == 0 else alt_fill()
            for ci, col in enumerate(cols_to_show, start=1):
                val = row_vals[col]
                cell = ws.cell(row=ri, column=ci)
                if not isinstance(val, str) and pd.isna(val):
                    cell.value = ""
                elif col in date_cols_set and hasattr(val, "date"):
                    cell.value = val.to_pydatetime()
                    cell.number_format = DATE_FMT
                else:
                    cell.value = val
                cell.fill = fill
                cell.font = body_font(10)
                cell.border = thin_border()
                cell.alignment = xleft() if ci <= 2 else xcenter()
            ws.row_dimensions[ri].height = 16
        autofit(ws, max_w=40, min_w=10)
        ws.freeze_panes = "A2"

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ═════════════════════════════════════════════════════════════════════════════

st.title("CAPA KPI Dashboard")
st.caption(f"Report date: {TODAY.strftime('%B %d, %Y')}")

uploaded_files = st.file_uploader(
    "Upload your export_CAPA .xls or .xlsx files",
    type=["xls", "xlsx"],
    accept_multiple_files=True,
    help="Select one or more `export_CAPA *.xls` or `export_CAPA *.xlsx` files exported from your CAPA system.",
)

if not uploaded_files:
    st.info("Upload one or more `export_CAPA *.xls` or `*.xlsx` files to get started.")
    st.stop()

all_capas, locations = load_data(uploaded_files)

# ── Sidebar ───────────────────────────────────────────────────────────────────
st.sidebar.header("Settings")

method = st.sidebar.radio(
    "Calculation method",
    options=["official", "taskdates"],
    format_func=lambda x: "Official (Date Closed)" if x == "official" else "Task Date Priority",
    help=(
        "**Official** uses only the 'Date closed' field on the Capas sheet.\n\n"
        "**Task Date Priority** uses the latest completed task date from the "
        "Taken sheet when available, falling back to the Capas Date closed."
    ),
)

selected_locations = st.sidebar.multiselect(
    "Filter by location",
    options=locations,
    default=locations,
)

if not selected_locations:
    st.warning("Select at least one location.")
    st.stop()

filtered = all_capas[all_capas["Location"].isin(selected_locations)].copy()
metrics = compute_metrics(filtered, method)

# ── KPI cards ─────────────────────────────────────────────────────────────────
method_label = "Official" if method == "official" else "Task Date Priority"
st.subheader(f"KPIs — {method_label}")

c1, c2, c3 = st.columns(3)
c1.metric("Avg Days to Close (2025)",
          f"{metrics['avg_2025']:.1f}" if pd.notna(metrics["avg_2025"]) else "N/A")
c2.metric("Total Closed in 2025", f"{metrics['closed_2025']:,}")
c3.metric("Avg Days to Close (2026)",
          f"{metrics['avg_2026']:.1f}" if pd.notna(metrics["avg_2026"]) else "N/A")

c4, c5, c6 = st.columns(3)
c4.metric("Total Closed in 2026", f"{metrics['closed_2026']:,}")
c5.metric("Currently Open", f"{metrics['open']:,}")
c6.metric(f"Open > {OPEN_THRESHOLD_DAYS} Days", f"{metrics['open_gt90']:,}")

# ----- Trend chart (average days to close per month) -----
trend_src = (
    filtered.dropna(subset=["Date closed", "Date of notification"])
    [["Date closed", "Date of notification"]]
    .copy()
)
if not trend_src.empty:
    trend_src["month"] = trend_src["Date closed"].dt.to_period("M")
    trend_src["days"] = (trend_src["Date closed"] - trend_src["Date of notification"]).dt.days
    trend_df = (
        trend_src.groupby("month", as_index=False)["days"].mean()
        .rename(columns={"days": "avg_days"})
    )
    trend_df["month"] = trend_df["month"].dt.to_timestamp()

    trend_chart = (
        alt.Chart(trend_df)
        .mark_line(point=True, color=f"#{MED_BLUE}")
        .encode(
            x=alt.X("month:T", title="Month"),
            y=alt.Y("avg_days:Q", title="Avg days to close"),
        )
        .properties(width=700, height=300, title="Close‑time trend")
    )
    st.altair_chart(trend_chart, use_container_width=True)
    st.caption("Average days from notification to closure, aggregated by month.")
else:
    st.info("Not enough closed-CAPA data for the trend chart.")

# ── Location table ────────────────────────────────────────────────────────────
st.subheader("Performance by Location")
st.dataframe(metrics["df_location"], use_container_width=True, hide_index=True)

# ── Detail tabs ───────────────────────────────────────────────────────────────
st.subheader("Detail Tables")
tab_names = list(metrics["details"].keys())
tabs = st.tabs(tab_names)
for tab, name in zip(tabs, tab_names):
    with tab:
        df = metrics["details"][name]
        st.dataframe(df, use_container_width=True, hide_index=True)
        st.caption(f"{len(df)} records")

# ── Download reports ──────────────────────────────────────────────────────────
st.divider()
st.subheader("Download Excel Reports")

dl1, dl2 = st.columns(2)

with dl1:
    official_metrics = compute_metrics(filtered, "official")
    official_bytes = build_excel_report(official_metrics, "official")
    st.download_button(
        label="Download Official Report",
        data=official_bytes,
        file_name="CAPA_Metrics_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

with dl2:
    taskdates_metrics = compute_metrics(filtered, "taskdates")
    taskdates_bytes = build_excel_report(taskdates_metrics, "taskdates")
    st.download_button(
        label="Download Task Dates Report",
        data=taskdates_bytes,
        file_name="CAPA_Metrics_Report_TaskDates.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
