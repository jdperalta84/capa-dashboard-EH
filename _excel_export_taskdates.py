from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import pandas as pd

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY        = "1F3864"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "DCE6F1"
ALT_ROW     = "EEF3FA"
WHITE       = "FFFFFF"
DARK_TEXT   = "1F3864"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D9D9D9"
ORANGE      = "C55A11"
GREEN       = "375623"
GREEN_BG    = "E2EFDA"
RED_BG      = "FCE4D6"

def navy_fill():    return PatternFill("solid", fgColor=NAVY)
def blue_fill():    return PatternFill("solid", fgColor=MED_BLUE)
def light_fill():   return PatternFill("solid", fgColor=LIGHT_BLUE)
def alt_fill():     return PatternFill("solid", fgColor=ALT_ROW)
def white_fill():   return PatternFill("solid", fgColor=WHITE)
def gray_fill():    return PatternFill("solid", fgColor=LIGHT_GRAY)

def hdr_font(sz=11):  return Font(name="Arial", bold=True, color=WHITE, size=sz)
def body_font(sz=11): return Font(name="Arial", size=sz, color=DARK_TEXT)
def bold_font(sz=11): return Font(name="Arial", bold=True, size=sz, color=DARK_TEXT)
def navy_font(sz=11): return Font(name="Arial", bold=True, size=sz, color=NAVY)

def thin_border():
    s = Side(style="thin", color=MED_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=False):   return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def right_align():      return Alignment(horizontal="right",  vertical="center")

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def autofit(ws, max_w=40, min_w=10):
    for col_cells in ws.columns:
        length = min_w
        for cell in col_cells:
            if cell.value:
                length = max(length, min(max_w, len(str(cell.value)) + 3))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill   = navy_fill()
        cell.font   = hdr_font()
        cell.border = thin_border()
        cell.alignment = center()

def style_data_rows(ws, start_row, end_row, ncols, start_col=1):
    for r in range(start_row, end_row + 1):
        fill = white_fill() if r % 2 == 0 else alt_fill()
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.fill   = fill
            cell.border = thin_border()
            cell.font   = body_font()
            cell.alignment = center() if c > start_col else left()

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ═══════════════════════════════════════════════════════════════════════════════
# 1. DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
dash = wb.create_sheet("Dashboard")
dash.sheet_view.showGridLines = False

# Title
dash.merge_cells("B2:I2")
dash["B2"].value     = "CAPA KPI Dashboard — Task Date Priority"
dash["B2"].font      = Font(name="Arial", bold=True, size=22, color=NAVY)
dash["B2"].alignment = left()

dash.merge_cells("B3:I3")
dash["B3"].value     = f"Report date: {TODAY.strftime('%B %d, %Y')}"
dash["B3"].font      = Font(name="Arial", size=11, color="595959")
dash["B3"].alignment = left()

dash.row_dimensions[2].height = 36
dash.row_dimensions[3].height = 18

# KPI card helper
def kpi_card(ws, label_row, val_row, col_start, col_end, label, value, bg_val=LIGHT_BLUE):
    col_s = get_column_letter(col_start)
    col_e = get_column_letter(col_end)
    # label band
    ws.merge_cells(f"{col_s}{label_row}:{col_e}{label_row}")
    lc = ws[f"{col_s}{label_row}"]
    lc.value     = label
    lc.fill      = navy_fill()
    lc.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    lc.alignment = center()
    ws.row_dimensions[label_row].height = 22
    # value band
    ws.merge_cells(f"{col_s}{val_row}:{col_e}{val_row}")
    vc = ws[f"{col_s}{val_row}"]
    vc.value     = value
    vc.fill      = PatternFill("solid", fgColor=bg_val)
    vc.font      = Font(name="Arial", bold=True, size=26, color=NAVY)
    vc.alignment = center()
    ws.row_dimensions[val_row].height = 44
    # border around full card
    for r in [label_row, val_row]:
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = thin_border()

# rows: cards start at row 5, 2 rows each + 1 gap
cards = [
    ("Avg Days to Close (2025)", f"{avg_days_to_close_2025:.1f} days"),
    ("Total Closed in 2025",     f"{total_closed_2025:,}"),
    ("Avg Days to Close (2026)", f"{avg_days_to_close_2026:.1f} days"),
    ("Total Closed in 2026",     f"{total_closed_2026:,}"),
    ("Currently Open",           f"{total_open:,}"),
    (f"Open > {OPEN_THRESHOLD_DAYS} Days",  f"{total_open_gt90:,}"),
]

card_positions = [
    (5, 6, 2, 5),   # row_label, row_val, col_start, col_end  — card 1
    (5, 6, 6, 9),   # card 2
    (8, 9, 2, 5),   # card 3
    (8, 9, 6, 9),   # card 4
    (11,12, 2, 5),  # card 5
    (11,12, 6, 9),  # card 6
]

# colour last two cards differently to signal attention
card_colors = [LIGHT_BLUE, LIGHT_BLUE, LIGHT_BLUE, LIGHT_BLUE, RED_BG, RED_BG]

for i, (label, value) in enumerate(cards):
    lr, vr, cs, ce = card_positions[i]
    kpi_card(dash, lr, vr, cs, ce, label, value, bg_val=card_colors[i])

# Location table
tbl_start = 15
dash.merge_cells(f"B{tbl_start}:H{tbl_start}")
dash[f"B{tbl_start}"].value     = "Performance by Location"
dash[f"B{tbl_start}"].font      = Font(name="Arial", bold=True, size=13, color=WHITE)
dash[f"B{tbl_start}"].fill      = navy_fill()
dash[f"B{tbl_start}"].alignment = left()
dash.row_dimensions[tbl_start].height = 22

loc_cols = ["Location", "Avg Days to Close (2025)", "Closed 2025",
            "Avg Days to Close (2026)", "Closed 2026", "Open", "Open > 90 days"]

hdr_row = tbl_start + 1
for ci, col in enumerate(loc_cols, start=2):
    cell = dash.cell(row=hdr_row, column=ci)
    cell.value     = col
    cell.fill      = blue_fill()
    cell.font      = hdr_font(10)
    cell.border    = thin_border()
    cell.alignment = center(wrap=True)
dash.row_dimensions[hdr_row].height = 30

for ri, (_, row_data) in enumerate(df_location.iterrows(), start=hdr_row + 1):
    fill = white_fill() if ri % 2 == 0 else alt_fill()
    values = [row_data["Location"],
              row_data["Avg Days to Close (2025)"],
              row_data["Closed 2025"],
              row_data["Avg Days to Close (2026)"],
              row_data["Closed 2026"],
              row_data["Open"],
              row_data["Open > 90 days"]]
    for ci, val in enumerate(values, start=2):
        cell = dash.cell(row=ri, column=ci)
        cell.value     = val
        cell.fill      = fill
        cell.font      = body_font(10)
        cell.border    = thin_border()
        cell.alignment = left() if ci == 2 else center()
    dash.row_dimensions[ri].height = 18

# Column widths on dashboard
col_widths = {2: 16, 3: 22, 4: 14, 5: 22, 6: 14, 7: 10, 8: 14, 9: 3}
for col, w in col_widths.items():
    set_col_width(dash, col, w)
set_col_width(dash, 1, 3)   # left margin

dash.freeze_panes = "B5"

# ═══════════════════════════════════════════════════════════════════════════════
# 2. BY LOCATION
# ═══════════════════════════════════════════════════════════════════════════════
loc_ws = wb.create_sheet("By Location")
loc_ws.sheet_view.showGridLines = False

loc_headers = ["Location", "Avg Days to Close (2025)", "Closed 2025",
               "Avg Days to Close (2026)", "Closed 2026", "Open", "Open > 90 Days"]

for ci, h in enumerate(loc_headers, start=1):
    cell = loc_ws.cell(row=1, column=ci)
    cell.value     = h
    cell.fill      = navy_fill()
    cell.font      = hdr_font()
    cell.border    = thin_border()
    cell.alignment = center(wrap=True)
loc_ws.row_dimensions[1].height = 30

for ri, (_, row_data) in enumerate(df_location.iterrows(), start=2):
    fill = white_fill() if ri % 2 == 0 else alt_fill()
    values = [row_data["Location"],
              row_data["Avg Days to Close (2025)"],
              row_data["Closed 2025"],
              row_data["Avg Days to Close (2026)"],
              row_data["Closed 2026"],
              row_data["Open"],
              row_data["Open > 90 days"]]
    for ci, val in enumerate(values, start=1):
        cell = loc_ws.cell(row=ri, column=ci)
        cell.value     = val
        cell.fill      = fill
        cell.font      = body_font()
        cell.border    = thin_border()
        cell.alignment = left() if ci == 1 else center()
    loc_ws.row_dimensions[ri].height = 18

autofit(loc_ws, max_w=28)
loc_ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# 3. SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
sum_ws = wb.create_sheet("Summary")
sum_ws.sheet_view.showGridLines = False

for ci, h in enumerate(["Metric", "Value"], start=1):
    cell = sum_ws.cell(row=1, column=ci)
    cell.value     = h
    cell.fill      = navy_fill()
    cell.font      = hdr_font()
    cell.border    = thin_border()
    cell.alignment = center()
sum_ws.row_dimensions[1].height = 24

summary_rows = [
    ("Avg Days to Close (2025)",  f"{avg_days_to_close_2025:.1f}"),
    ("Total Closed in 2025",      total_closed_2025),
    ("Avg Days to Close (2026)",  f"{avg_days_to_close_2026:.1f}"),
    ("Total Closed in 2026",      total_closed_2026),
    ("Currently Open",            total_open),
    (f"Open > {OPEN_THRESHOLD_DAYS} Days", total_open_gt90),
]

for ri, (metric, value) in enumerate(summary_rows, start=2):
    fill = white_fill() if ri % 2 == 0 else alt_fill()
    for ci, val in enumerate([metric, value], start=1):
        cell = sum_ws.cell(row=ri, column=ci)
        cell.value     = val
        cell.fill      = fill
        cell.border    = thin_border()
        cell.alignment = left() if ci == 1 else center()
        cell.font      = body_font() if ci == 1 else bold_font()
    sum_ws.row_dimensions[ri].height = 20

sum_ws.column_dimensions["A"].width = 36
sum_ws.column_dimensions["B"].width = 16
sum_ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# 4. LOGIC NOTES
# ═══════════════════════════════════════════════════════════════════════════════
log_ws = wb.create_sheet("Logic Notes")
log_ws.sheet_view.showGridLines = False

# Title
log_ws.merge_cells("A1:B1")
log_ws["A1"].value     = "CAPA Metrics — Logic & Methodology"
log_ws["A1"].font      = Font(name="Arial", bold=True, size=16, color=WHITE)
log_ws["A1"].fill      = navy_fill()
log_ws["A1"].alignment = center()
log_ws.row_dimensions[1].height = 32

LOGIC_CONTENT = [
    # (rule, detail, is_section)
    ("ALTERNATIVE METHOD: TASK DATE PRIORITY", "", True),
    ("How this differs from the official method",
     "The official report uses only the 'Date closed' field on the main Capas tab. This version prioritises task completion dates from the Taken sheet, falling back to the Capas date if no completed tasks exist.", False),
    ("Why task dates may be considered",
     "Task completion dates reflect when the actual corrective work was finished. In some cases the Capas Date closed is entered later as an administrative step, meaning the official date may overstate how long resolution took.", False),
    ("Known trade-off: shorter averages",
     "Using task dates produces lower avg days figures. In our Portugal validation, the avg dropped from 49.9 days (official) to 23.3 days (task date) - a 53% reduction. Whether this is more accurate depends on your definition of closed.", False),
    ("Known trade-off: more CAPAs counted as closed",
     "CAPAs with completed tasks but no Date closed on the Capas tab are counted as closed here. This increases closed counts and reduces open counts vs. the official method.", False),
    ("Subjectivity risk",
     "This method requires deciding how to handle partial task lists. Here, any completed task with a date is sufficient to derive a closed date (the max completion date). This assumption may not always be appropriate.", False),

    ("HOW WE DETERMINE IF A CAPA IS CLOSED (THIS VERSION)", "", True),
    ("Priority 1 - Task dates",
     "If any tasks in the Taken sheet are marked Completed = Yes with a Date of completion, the CAPA is closed. Closed date = latest of those completion dates.", False),
    ("Priority 2 - Capas sheet fallback",
     "If no completed tasks with dates exist, use the Date closed field on the main Capas tab.", False),
    ("Priority 3 - Open",
     "If neither source has a date, the CAPA is open.", False),
    ("Status column",
     "NOT used. Status values like Afgesloten are ignored.", False),

    ("HOW EACH METRIC IS CALCULATED", "", True),
    ("Avg days to close (2025)",
     "All CAPAs with an effective closed date in 2025. Formula: effective closed date - Date of notification, averaged.", False),
    ("Avg days to close (2026)",
     "Same logic applied to CAPAs with an effective closed date in 2026.", False),
    ("Total closed in 2025 / 2026",
     "Count of CAPAs where the year of the effective closed date matches.", False),
    ("Currently open",
     "CAPAs with no effective closed date from either source.", False),
    ("Open > 90 days",
     "Open CAPAs where: today - Date of notification > 90 days.", False),

    ("DATA SOURCES", "", True),
    ("Input files",
     "All files matching 'export_CAPA *.xls' in the same folder as the script.", False),
    ("Sheets used",
     "'Capas' sheet: main CAPA records. 'Taken' sheet: action tasks used to derive effective closed dates.", False),
    ("To refresh",
     "Replace the export_CAPA *.xls files and run:  py capa_metrics_taskdates.py", False),
]

row = 2
for rule, detail, is_section in LOGIC_CONTENT:
    if is_section:
        log_ws.merge_cells(f"A{row}:B{row}")
        cell = log_ws[f"A{row}"]
        cell.value     = rule
        cell.fill      = blue_fill()
        cell.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
        cell.alignment = left()
        cell.border    = thin_border()
        log_ws.row_dimensions[row].height = 22
    else:
        fill = gray_fill() if row % 2 == 0 else white_fill()
        for ci, val in enumerate([rule, detail], start=1):
            cell = log_ws.cell(row=row, column=ci)
            cell.value     = val
            cell.fill      = fill
            cell.font      = bold_font(10) if ci == 1 else body_font(10)
            cell.alignment = left(wrap=True)
            cell.border    = thin_border()
        log_ws.row_dimensions[row].height = 42
    row += 1

log_ws.column_dimensions["A"].width = 45
log_ws.column_dimensions["B"].width = 80

# ═══════════════════════════════════════════════════════════════════════════════
# 5–8. DETAIL SHEETS
# ═══════════════════════════════════════════════════════════════════════════════
cols_to_show = ["Location", "Number", "Date of notification",
                "Date closed", "Effective closed date", "Status"]

detail_sheets = [
    ("Closed 2025 Detail", all_capas[closed_2025]),
    ("Closed 2026 Detail", all_capas[closed_2026]),
    ("Open Detail",        all_capas[is_open]),
    ("Open >90d Detail",   all_capas[open_gt90]),
]

DATE_FMT = "YYYY-MM-DD"
date_cols = {"Date of notification", "Date closed", "Effective closed date"}

for sheet_name, df_detail in detail_sheets:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Header
    for ci, col in enumerate(cols_to_show, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value     = col
        cell.fill      = navy_fill()
        cell.font      = hdr_font()
        cell.border    = thin_border()
        cell.alignment = center(wrap=True)
    ws.row_dimensions[1].height = 24

    # Data
    df_out = df_detail[cols_to_show].copy()
    for ri, (_, row_vals) in enumerate(df_out.iterrows(), start=2):
        fill = white_fill() if ri % 2 == 0 else alt_fill()
        for ci, col in enumerate(cols_to_show, start=1):
            val = row_vals[col]
            cell = ws.cell(row=ri, column=ci)
            if pd.isna(val) if not isinstance(val, str) else False:
                cell.value = ""
            elif col in date_cols and hasattr(val, "date"):
                cell.value      = val.to_pydatetime()
                cell.number_format = DATE_FMT
            else:
                cell.value = val
            cell.fill      = fill
            cell.font      = body_font(10)
            cell.border    = thin_border()
            cell.alignment = left() if ci <= 2 else center()
        ws.row_dimensions[ri].height = 16

    autofit(ws, max_w=40, min_w=10)
    ws.freeze_panes = "A2"

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
output_path = os.path.join(SCRIPT_DIR, "CAPA_Metrics_Report_TaskDates.xlsx")
wb.save(output_path)
print(f"  Report saved to: {output_path}")
